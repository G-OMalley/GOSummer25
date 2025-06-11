# create_lookyloo_report.py
# A standalone script to generate a formatted "LookyLoo" Excel report.

import pandas as pd
import os
from pathlib import Path
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
import numpy as np
import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_DIR = Path(__file__).resolve().parent 
INFO_DIR = BASE_DIR.parent / "INFO" 
OUTPUT_DIR = BASE_DIR / "MarketAnalysis_Report_Output" # MODIFIED: Output will be in a subdirectory

# --- File Paths ---
PRICES_FILE = INFO_DIR / "PRICES.csv"
ICE_FILE = INFO_DIR / "ICE_daily.xlsx"
HISTORICAL_FOM_FILE = INFO_DIR / "HistoricalFOM.csv"

# --- Report Parameters ---
TODAY = date.today()
LOOKBACK_YEARS = 5 
HENRY_HUB_NAME = 'Henry'
COMPLETENESS_LOOKBACK_DAYS = 90
COMPLETENESS_THRESHOLD = 0.75

# --- Data Loading Functions ---

def robust_date_parser(df, date_col_name='Date'):
    """
    Parses a DataFrame column that may contain mixed date formats (strings and Excel serial numbers).
    """
    parsed_dates = pd.to_datetime(df[date_col_name], errors='coerce')
    failed_indices = parsed_dates[parsed_dates.isna()].index
    
    if not failed_indices.empty:
        serial_dates = pd.to_numeric(df.loc[failed_indices, date_col_name], errors='coerce')
        excel_dates = pd.to_datetime(serial_dates, origin='1899-12-30', unit='D', errors='coerce')
        parsed_dates.fillna(excel_dates, inplace=True)
        
    return parsed_dates

def load_data_sources():
    """Loads all necessary data files into pandas DataFrames."""
    print("Loading data sources...")
    try:
        df_prices = pd.read_csv(PRICES_FILE)
        df_prices['Date'] = robust_date_parser(df_prices, 'Date').dt.date
        df_prices.dropna(subset=['Date'], inplace=True) 
        print("PRICES.csv loaded and dates parsed successfully.")
    except Exception as e:
        print(f"FATAL: Could not load or parse PRICES.csv. Error: {e}")
        return None, None, None

    try:
        df_ice = pd.read_excel(
            ICE_FILE, 
            sheet_name=0, 
            header=None, 
            skiprows=3,
            usecols=[1, 34, 35]
        )
        df_ice.columns = ['Component', 'Mark0', 'Price Region']
        print("ICE_daily.xlsx loaded and structured correctly.")
    except Exception as e:
        print(f"FATAL: Could not load ICE_daily.xlsx. Error: {e}")
        return None, None, None

    try:
        df_fom = pd.read_csv(HISTORICAL_FOM_FILE)
        print("HistoricalFOM.csv loaded.")
    except Exception as e:
        print(f"FATAL: Could not load HistoricalFOM.csv. Error: {e}")
        return None, None, None

    return df_prices, df_ice, df_fom

def get_active_components(df_prices, current_date):
    """Filters for active market components based on data completeness."""
    print("Filtering for active market components...")
    end_date = current_date
    start_date = end_date - timedelta(days=COMPLETENESS_LOOKBACK_DAYS - 1)
    
    active_components = []
    try:
        break_col_index = df_prices.columns.to_list().index('Unnamed: 58')
        component_candidates = df_prices.columns[1:break_col_index].tolist()
    except ValueError:
        component_candidates = [col for col in df_prices.columns if col != 'Date']

    for component in component_candidates:
        if component.lower() == HENRY_HUB_NAME.lower(): continue
            
        series_in_period = df_prices[(df_prices['Date'] >= start_date) & (df_prices['Date'] <= end_date)][component]
        if series_in_period.empty: continue
        
        completeness_ratio = series_in_period.notna().sum() / COMPLETENESS_LOOKBACK_DAYS
        if completeness_ratio >= COMPLETENESS_THRESHOLD:
            active_components.append(component)

    print(f"Found {len(active_components)} active gas components.")
    return active_components

# --- Calculation Functions ---

def get_fom_basis(df_fom, component, target_month, target_year):
    """Gets a single First of Month basis value from the HistoricalFOM DataFrame."""
    month_name = date(target_year, target_month, 1).strftime('%B')
    val = df_fom[
        (df_fom['market_component'] == component) &
        (df_fom['settlement_month'] == month_name) &
        (df_fom['settlement_year'] == target_year)
    ]['settlement_basis']
    return val.iloc[0] if not val.empty else np.nan

def get_monthly_avg_basis(df_prices, component, target_month, target_year, is_current_month):
    """Calculates the average daily basis for a given month and year."""
    month_data = df_prices[
        (df_prices['Date'].apply(lambda d: d.month) == target_month) &
        (df_prices['Date'].apply(lambda d: d.year) == target_year)
    ]
    
    if is_current_month:
        month_data = month_data[month_data['Date'] <= TODAY]

    if month_data.empty or HENRY_HUB_NAME not in month_data.columns or component not in month_data.columns:
        return np.nan
        
    basis = pd.to_numeric(month_data[component], errors='coerce') - pd.to_numeric(month_data[HENRY_HUB_NAME], errors='coerce')
    return basis.mean()


# --- Main Logic ---
def main():
    """Main function to generate the LookyLoo report."""
    df_prices, df_ice, df_fom = load_data_sources()
    if df_prices is None or df_ice is None or df_fom is None:
        return

    gas_components = get_active_components(df_prices, TODAY)
    
    region_map = df_ice.set_index('Component')['Price Region'].to_dict()
    fwd_mark_map_prompt = df_ice.set_index('Component')['Mark0'].to_dict()
    
    report_data = []
    
    current_month_date = TODAY
    prompt_month_date = TODAY + relativedelta(months=1)
    prompt2_month_date = TODAY + relativedelta(months=2)

    for component in sorted(gas_components):
        print(f"Processing component: {component}")
        row = {'Market Component': component}
        row['Price Region'] = region_map.get(component, 'N/A')
        
        # Current Month History
        for i in range(LOOKBACK_YEARS, 0, -1):
            hist_date = current_month_date - relativedelta(years=i)
            col_name_fom = f"{hist_date.strftime('%b%y')} FoM"
            col_name_avg = f"{hist_date.strftime('%b%y')} $"
            row[col_name_fom] = get_fom_basis(df_fom, component, hist_date.month, hist_date.year)
            row[col_name_avg] = get_monthly_avg_basis(df_prices, component, hist_date.month, hist_date.year, False)

        # Current Month Actuals
        col_name_fom = f"{current_month_date.strftime('%b%y')} FoM"
        col_name_avg = f"{current_month_date.strftime('%b%y')} $"
        row[col_name_fom] = get_fom_basis(df_fom, component, current_month_date.month, current_month_date.year)
        row[col_name_avg] = get_monthly_avg_basis(df_prices, component, current_month_date.month, current_month_date.year, True)
        
        # Prior 7 Day Average
        seven_days_ago = TODAY - timedelta(days=7)
        seven_day_data = df_prices[(df_prices['Date'] > seven_days_ago) & (df_prices['Date'] <= TODAY)]
        if not seven_day_data.empty and component in seven_day_data.columns:
            basis_7d = pd.to_numeric(seven_day_data[component], errors='coerce') - pd.to_numeric(seven_day_data[HENRY_HUB_NAME], errors='coerce')
            row['7d Avg'] = basis_7d.mean()
        else:
            row['7d Avg'] = np.nan
        
        # Prompt Month History
        for i in range(LOOKBACK_YEARS, 0, -1):
            hist_date = prompt_month_date - relativedelta(years=i)
            col_name_fom = f"{hist_date.strftime('%b%y')} FoM"
            col_name_avg = f"{hist_date.strftime('%b%y')} $"
            row[col_name_fom] = get_fom_basis(df_fom, component, hist_date.month, hist_date.year)
            row[col_name_avg] = get_monthly_avg_basis(df_prices, component, hist_date.month, hist_date.year, False)
            
        # Forward Mark (Prompt)
        row['Forward Mark'] = fwd_mark_map_prompt.get(component, np.nan)
            
        # Prompt + 2 Month History
        for i in range(LOOKBACK_YEARS, 0, -1):
            hist_date = prompt2_month_date - relativedelta(years=i)
            col_name_fom = f"{hist_date.strftime('%b%y')} FoM"
            col_name_avg = f"{hist_date.strftime('%b%y')} $"
            row[col_name_fom] = get_fom_basis(df_fom, component, hist_date.month, hist_date.year)
            row[col_name_avg] = get_monthly_avg_basis(df_prices, component, hist_date.month, hist_date.year, False)
            
        # Final Calculation Columns
        row['Fwd vs 7d'] = row['Forward Mark'] - row['7d Avg']
        
        prior_year_prompt_avg_col = (prompt_month_date - relativedelta(years=1)).strftime('%b%y') + " $"
        row['Fwd vs PY'] = row['Forward Mark'] - row.get(prior_year_prompt_avg_col, np.nan)

        report_data.append(row)

    df_report = pd.DataFrame(report_data)
    
    print("Generating styled Excel report...")
    
    # --- Excel Styling and Generation ---
    output_file_path = OUTPUT_DIR / "LookyLoo_Report.xlsx"
    writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
    
    df_report.to_excel(writer, sheet_name='LookyLoo', index=False)
    
    workbook  = writer.book
    worksheet = writer.sheets['LookyLoo']
    
    # --- Define openpyxl Style Objects ---
    highlight_fill = PatternFill(start_color='E0F2FF', end_color='E0F2FF', fill_type='solid')
    red_font = Font(color="9C0006")
    green_font = Font(color="006100")
    dg_fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    dg_font = Font(color="FFFFFF")
    mg_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    lg_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    dr_fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
    dr_font = Font(color="FFFFFF")
    mr_fill = PatternFill(start_color='F44336', end_color='F44336', fill_type='solid')
    lr_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

    header = df_report.columns.to_list()
    
    # Loop through all data cells to apply formatting
    for r_idx, row_data in enumerate(df_report.itertuples(index=False), 2): # r_idx starts at 2 for Excel rows
        for c_idx, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            col_name = header[c_idx - 1]
            
            # Apply base number formatting to all numeric cells
            if isinstance(cell_value, (int, float, np.number)):
                cell.number_format = '0.000'

            # Highlight '7d Avg' and 'Forward Mark' columns
            if col_name in ['7d Avg', 'Forward Mark']:
                cell.fill = highlight_fill
            
            # Color font for 'Fwd vs' columns
            if col_name in ['Fwd vs 7d', 'Fwd vs PY']:
                if pd.notna(cell_value):
                    cell.font = red_font if cell_value < 0 else green_font
            
            # Color background for '$' average columns
            match = re.match(r"(\w{3}\d{2}) \$", col_name)
            if match:
                month_year_prefix = match.group(1)
                fom_col_name = f"{month_year_prefix} FoM"
                if fom_col_name in df_report.columns:
                    fom_val = df_report.loc[r_idx - 2, fom_col_name] # get value from original df
                    avg_val = cell_value

                    if pd.notna(fom_val) and pd.notna(avg_val):
                        diff = avg_val - fom_val
                        if diff > 0.10: 
                            cell.fill = dg_fill
                            cell.font = dg_font
                        elif diff > 0.05: cell.fill = mg_fill
                        elif diff > 0.02: cell.fill = lg_fill
                        elif diff < -0.10: 
                            cell.fill = dr_fill
                            cell.font = dr_font
                        elif diff < -0.05: cell.fill = mr_fill
                        elif diff < -0.02: cell.fill = lr_fill

    # Add AutoFilter and auto-fit columns
    worksheet.auto_filter.ref = worksheet.dimensions
    for i, col in enumerate(df_report.columns, 1):
        column_letter = get_column_letter(i)
        column_len = max(df_report[col].astype(str).map(len).max(), len(col)) + 2 # Add padding
        worksheet.column_dimensions[column_letter].width = column_len

    writer.close()
    print(f"\nSuccessfully generated formatted Excel report: {output_file_path.resolve()}")

if __name__ == '__main__':
    main()